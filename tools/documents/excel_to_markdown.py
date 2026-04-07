import os
from typing import Optional, Callable
from langchain_core.tools import tool


@tool(description="将 Excel/CSV 转为 Markdown（无合并单元格）或 HTML 表格（包含合并单元格），并上传返回文件 URL")
def excel_to_markdown(file_path: str) -> str:
    """
    工具：将表格文件转换为 Markdown 并上传（可由模型直接调用）。

    功能概述：
    - 接受本地文件路径或远程 URL（以 http:// 或 https:// 开头）。
    - 支持输入格式：`.xlsx`, `.xls`, `.csv`。CSV/XLS 会先转换为 XLSX 再处理。
    - 对于不包含合并单元格的 sheet：输出标准 Markdown 表格（pandas.to_markdown），并将所有缺失值（NaN/None/NaT）替换为空字符串（""）。
    - 对于包含合并单元格的 sheet：输出 HTML 表格（<table>），并保留合并信息（在左上角单元格添加 `rowspan` / `colspan`），其余被合并的单元格会被跳过，从而在渲染器中保留合并效果。
    - 所有单元格在输出前都会安全转换为字符串；HTML 模式下会对内容做 HTML 转义并把换行替换为 `<br>`。

    参数：
    - file_path (str): 本地绝对路径或远程文件 URL。

    返回：
    - str: 成功返回上传后文件的访问 URL；失败返回可读的错误信息字符串。

    使用说明（模型作为工具调用）：
    - 仅需传入 `file_path`，工具会执行转换并上传，返回 URL 或错误字符串。
    - 若需要结构化输出（例如 JSON 包含 success/data/error），可在上层封装此工具或请求我修改为返回 JSON 字符串。
    - 为方便单元测试或本地调试，内部实现支持将上传函数作为参数注入（`upload_func`），但模型直接调用时不需要提供该参数。
    """
    tmp_file_path = None
    if file_path.startswith("http://") or file_path.startswith("https://"):
        import requests
        import tempfile

        try:
            response = requests.get(file_path)
            response.raise_for_status()

            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_path)[1]) as tmp_file:
                tmp_file.write(response.content)
                tmp_file_path = tmp_file.name

            file_path = tmp_file_path
        except Exception as e:
            print(f"Error downloading file from URL: {str(e)}")
            return f"Failed to download the file: {str(file_path)}"

    try:
        from tools.third_party.aliyun_oss_uploader import upload_file

        return process(file_path, upload_func=upload_file)
    except Exception as e:
        print(f"Error converting file to markdown: {str(e)}")
        return f"Failed to convert the file: {str(file_path)}"
    finally:
        if tmp_file_path and os.path.exists(tmp_file_path):
            os.remove(tmp_file_path)


def process(input_path: str, upload_func: Optional[Callable[[bytes, Optional[str]], str]] = None) -> str:
    import pandas as pd
    import time
    import openpyxl

    ext = os.path.splitext(input_path)[1].lower()
    cleanup_tmp = False
    if ext == '.csv':
        print("converting csv to xlsx...")
        xlsx_path = convert_csv_to_xlsx(input_path)
        cleanup_tmp = True
    elif ext == '.xls':
        print("converting xls to xlsx...")
        xlsx_path = convert_xls_to_xlsx(input_path)
        cleanup_tmp = True
    elif ext == '.xlsx':
        xlsx_path = input_path
    else:
        raise ValueError('Unsupported input type: ' + ext)

    try:
        print("processing xlsx...")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        markdown_content = ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            merged_cells_ranges = list(ws.merged_cells.ranges)

            if len(merged_cells_ranges) == 0:
                # --- 原始逻辑（无合并单元格） ---
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))

                if not data:
                    continue

                df = pd.DataFrame(data)

                if df.shape[0] >= 1:
                    raw_header = df.iloc[0].tolist()
                    header = [("" if pd.isna(h) else str(h)) for h in raw_header]
                    df = df.iloc[1:].reset_index(drop=True)
                    df.columns = header
                else:
                    continue

                df = df.where(pd.notna(df), "")

                rows_as_str = [[
                    ("" if (cell is None or (isinstance(cell, float) and pd.isna(cell)) or cell == "") else
                     (cell.decode('utf-8', errors='ignore') if isinstance(cell, (bytes, bytearray)) else str(cell))
                     ) for cell in row]
                    for row in df.values.tolist()
                ]
                df = pd.DataFrame(rows_as_str, columns=df.columns)

                sheet_md = f"## Sheet: {sheet_name}\n\n"
                sheet_md += df.to_markdown(index=False)
                markdown_content += sheet_md + "\n\n"
            else:
                # --- HTML 表格逻辑（有合并单元格） ---
                skip_cells = set()
                merge_attrs = {}

                for merged_range in merged_cells_ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    colspan = max_col - min_col + 1
                    rowspan = max_row - min_row + 1
                    merge_attrs[(min_row, min_col)] = {'rowspan': rowspan, 'colspan': colspan}
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            if r == min_row and c == min_col:
                                continue
                            skip_cells.add((r, c))

                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))

                if not data:
                    continue

                import html

                html_lines = ["<table>"]
                for r_idx, row in enumerate(data, start=1):
                    html_lines.append("  <tr>")
                    for c_idx, cell_value in enumerate(row, start=1):
                        if (r_idx, c_idx) in skip_cells:
                            continue

                        attrs = ""
                        if (r_idx, c_idx) in merge_attrs:
                            m = merge_attrs[(r_idx, c_idx)]
                            if m['rowspan'] > 1:
                                attrs += f' rowspan="{m["rowspan"]}"'
                            if m['colspan'] > 1:
                                attrs += f' colspan="{m["colspan"]}"'

                        # 安全地转换为字符串
                        if cell_value is None or (isinstance(cell_value, float) and pd.isna(cell_value)):
                            val_str = ""
                        elif isinstance(cell_value, (bytes, bytearray)):
                            val_str = cell_value.decode('utf-8', errors='ignore')
                        else:
                            val_str = str(cell_value)

                        # 转义 HTML 实体，并替换换行符
                        val_str = html.escape(val_str).replace("\n", "<br>")

                        tag = "th" if r_idx == 1 else "td"
                        html_lines.append(f"    <{tag}{attrs}>{val_str}</{tag}>")
                    html_lines.append("  </tr>")
                html_lines.append("</table>")

                sheet_md = f"## Sheet: {sheet_name}\n\n"
                sheet_md += "\n".join(html_lines)
                markdown_content += sheet_md + "\n\n"

        # Use a timestamped filename to avoid collisions when uploading multiple times
        filename = f"output_{int(time.time())}.md"
        md_file_url = upload_func(markdown_content.encode("utf-8"), filename)
        return md_file_url
    finally:
        if cleanup_tmp:
            try:
                os.remove(xlsx_path)
            except Exception:
                pass


def convert_csv_to_xlsx(csv_path: str) -> str:
    import pandas as pd
    import tempfile

    df = pd.read_csv(csv_path)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        df.to_excel(tmp_file.name, index=False)
        return tmp_file.name


def convert_xls_to_xlsx(xls_path: str) -> str:
    import pandas as pd
    import tempfile

    df = pd.read_excel(xls_path, engine='xlrd')
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        df.to_excel(tmp_file.name, index=False)
        return tmp_file.name
